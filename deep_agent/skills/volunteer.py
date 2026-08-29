import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from models.volunteer import Volunteer


EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "volunteerdata", "volunteers.xlsx")

HEADERS = [
    "Full Name", "Preferred Name", "Age", "Gender",
    "City / Area", "Email ID", "Mobile Number",
    "Why Volunteer", "Areas of Interest",
    "Skills / Expertise", "Previous Experience", "Previous Organization",
    "Preferred Days", "Preferred Time", "Hours Per Week",
    "Volunteering Mode", "Preferred Location",
    # Emergency & Safety
    "Emergency Contact Name", "Emergency Contact Relationship",
    "Emergency Contact Number", "Safety Information",
    # Consent
    "Consent: Code of Conduct", "Consent: Safeguarding",
    "Consent: Contact", "Consent: Photography", "Consent: Accuracy",
    # CPP Training
    "CPP Training Completed", "CPP Training Date",
    "Registered At",
]

EMAIL_COL_INDEX = 5      # 0-based index of Email ID column
CPP_COL_INDEX   = 26     # 0-based index of CPP Training Completed
CPP_DATE_INDEX  = 27     # 0-based index of CPP Training Date


def _ensure_workbook() -> None:
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Volunteers"
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)


def search_volunteer(email: str) -> Volunteer | None:
    # ── XLSX is the registration store — check it first ──────────────────────
    # Google Sheet is CPP quiz results only — checked separately in LoginHandler
    _ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[EMAIL_COL_INDEX] and row[EMAIL_COL_INDEX].lower() == email.lower():
            c = [str(v) if v else "" for v in row]
            vol = Volunteer()
            vol.full_name                      = c[0]
            vol.preferred_name                 = c[1]
            vol.age                            = c[2]
            vol.gender                         = c[3]
            vol.city_area                      = c[4]
            vol.email                          = c[5]
            vol.mobile_number                  = c[6]
            vol.why_volunteer                  = c[7]
            vol.areas_of_interest              = c[8]
            vol.skills_expertise               = c[9]
            vol.previous_experience            = c[10]
            vol.previous_organization          = c[11]
            vol.preferred_days                 = c[12]
            vol.preferred_time                 = c[13]
            vol.hours_per_week                 = c[14]
            vol.volunteering_mode              = c[15]
            vol.preferred_location             = c[16]
            vol.emergency_contact_name         = c[17]
            vol.emergency_contact_relationship = c[18]
            vol.emergency_contact_number       = c[19]
            vol.safety_information             = c[20]
            vol.consent_code_of_conduct        = c[21]
            vol.consent_safeguarding           = c[22]
            vol.consent_contact                = c[23]
            vol.consent_photography            = c[24]
            vol.consent_accuracy               = c[25]
            vol.cpp_training_completed         = c[26] if len(c) > 26 else "No"
            vol.cpp_training_date              = c[27] if len(c) > 27 else ""
            return vol
    return None


def create_volunteer(v: Volunteer) -> str:
    _ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        v.full_name, v.preferred_name, v.age, v.gender,
        v.city_area, v.email, v.mobile_number,
        v.why_volunteer, v.areas_of_interest,
        v.skills_expertise, v.previous_experience, v.previous_organization,
        v.preferred_days, v.preferred_time, v.hours_per_week,
        v.volunteering_mode, v.preferred_location,
        v.emergency_contact_name, v.emergency_contact_relationship,
        v.emergency_contact_number, v.safety_information,
        v.consent_code_of_conduct, v.consent_safeguarding,
        v.consent_contact, v.consent_photography, v.consent_accuracy,
        "No", "",                                          # cpp_training_completed, cpp_training_date
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])
    wb.save(EXCEL_FILE)
    vol_id = f"VOL-{ws.max_row - 1:04d}"
    return vol_id


def mark_cpp_complete(email: str) -> bool:
    _ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[EMAIL_COL_INDEX].value and row[EMAIL_COL_INDEX].value.lower() == email.lower():
            row[CPP_COL_INDEX].value  = "Yes"
            row[CPP_DATE_INDEX].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            wb.save(EXCEL_FILE)
            return True
    return False
